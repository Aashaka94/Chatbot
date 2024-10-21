import openai
import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Set up OpenAI API key
openai.api_key = ''

# Function to handle Medication Adherence % and show changes as Mapped from "X" to "Y"
def map_medication_adherence(adherence):
    adherence_str = str(adherence).strip().lower()  # Ensure adherence is a string

    if adherence_str in ['high', 'medium', 'low']:
        return adherence_str.capitalize(), None, False  # No change
    
    try:
        adherence_value = float(adherence_str.replace('%', '').strip())
        
        if adherence_value < 0 or adherence_value > 100:
            return 'Outlier', 'Medication Adherence: Mapped from "{}" to "Outlier"'.format(adherence), True
        elif adherence_value >= 75:
            return 'High', 'Medication Adherence: Mapped from "{}" to "High"'.format(adherence), True
        elif 50 <= adherence_value < 75:
            return 'Medium', 'Medication Adherence: Mapped from "{}" to "Medium"'.format(adherence), True
        else:
            return 'Low', 'Medication Adherence: Mapped from "{}" to "Low"'.format(adherence), True
    except ValueError:
        return 'Invalid', 'Medication Adherence: Mapped from "{}" to "Invalid"'.format(adherence), True

# Function to extract mapped values from GenAI response
def extract_mapped_value(response_text):
    match = re.findall(r"'(.+?)'", response_text)
    if match and len(match) > 1:
        return match[1]  # Extract the second item as the mapped value
    return None

# Function to clean the HCP ID by removing non-digit characters
def clean_hcp_id(hcp_id):
    cleaned_hcp_id = re.sub(r'\D', '', str(hcp_id))  # Ensure hcp_id is a string and remove non-digit characters
    return cleaned_hcp_id

# Function to use GenAI to find the closest match for HCP ID (strictly 5 digits)
def map_hcp_id(hcp_id, veeva_hcp_ids):
    veeva_hcp_ids_5digit = [vid for vid in veeva_hcp_ids if len(vid) == 5]  # Filter to only 5-digit HCP IDs

    # Clean the HCP ID to extract only digits
    cleaned_hcp_id = clean_hcp_id(hcp_id)

    # Check if the cleaned HCP ID is a valid 5-digit ID in the Veeva list
    if cleaned_hcp_id in veeva_hcp_ids_5digit:
        if cleaned_hcp_id != str(hcp_id):  # If there was a cleaning operation
            return cleaned_hcp_id, 'HCP ID: Mapped from "{}" to "{}"'.format(hcp_id, cleaned_hcp_id), True
        else:
            return cleaned_hcp_id, None, False  # No change needed if no cleaning occurred

    try:
        # GenAI will map the closest valid 5-digit HCP ID from the Veeva list
        prompt = f"Map the HCP ID '{cleaned_hcp_id}' to the closest valid 5-digit HCP ID from the list: {veeva_hcp_ids_5digit}."
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that finds and corrects the closest 5-digit HCP IDs."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=150,
            temperature=0.5
        )
        corrected_hcp_id_full = response['choices'][0]['message']['content'].strip()
        corrected_hcp_id = extract_mapped_value(corrected_hcp_id_full)
        
        # If GenAI finds a valid match and it's different from the cleaned ID, apply the mapping
        if corrected_hcp_id and corrected_hcp_id != cleaned_hcp_id:
            return corrected_hcp_id, 'HCP ID: Mapped from "{}" to "{}"'.format(hcp_id, corrected_hcp_id), True
        else:
            return cleaned_hcp_id, 'HCP ID: Corrected "{}" but no valid 5-digit match found'.format(hcp_id), True
    except Exception as e:
        return cleaned_hcp_id, 'HCP ID: Error mapping "{}"'.format(hcp_id), True

# Function to use GenAI for mapping drug names
def map_drug_name(drug_name, veeva_product_names):
    if drug_name in veeva_product_names:
        return drug_name, None, False  # No change needed for valid Drug Name

    try:
        prompt = f"Map the drug name '{drug_name}' to the most similar one from the following list: {veeva_product_names}"
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that maps drug names."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=100,
            temperature=0.5
        )
        mapped_drug_full = response['choices'][0]['message']['content'].strip()
        mapped_drug = extract_mapped_value(mapped_drug_full)
        if mapped_drug and mapped_drug != drug_name:
            return mapped_drug, 'Drug Name: Mapped from "{}" to "{}"'.format(drug_name, mapped_drug), True
        else:
            return drug_name, None, False
    except Exception as e:
        return drug_name, None, False

# Function to save to Excel with highlighting for changes
def save_to_excel_with_highlight(df, file_name, changed_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapped Data"
    
    # Define light green fill for changed cells
    highlight_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    # Write headers
    for col_num, column_title in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write data and highlight changes
    for row_num, row_data in df.iterrows():
        for col_num, (col_name, cell_value) in enumerate(row_data.items(), 1):
            cell = ws.cell(row=row_num + 2, column=col_num, value=cell_value)
            # If this row/column cell has changed, highlight it
            if changed_rows[row_num].get(col_name, False):
                cell.fill = highlight_fill

    wb.save(file_name)
    print(f"Data saved to {file_name} with highlights for changed cells.")

# Main semantic mapping function
def semantic_mapping(veeva_df, xponent_df, output_file_name):
    try:
        # Ensure all necessary fields in Veeva are treated as strings
        veeva_df['HCP ID'] = veeva_df['HCP ID'].astype(str)
        veeva_df['Product Name'] = veeva_df['Product Name'].astype(str)
        veeva_df['Customer Group'] = veeva_df['Customer Group'].astype(str)

        # Unique lists from Veeva for matching
        veeva_hcp_ids = veeva_df['HCP ID'].unique().tolist()
        veeva_product_names = veeva_df['Product Name'].tolist()
        veeva_hcp_to_gpo = veeva_df.set_index('HCP ID')['Customer Group'].to_dict()

        xponent_cleaned_data = []
        changed_rows = []
        operations_performed = []

        for index, row in xponent_df.iterrows():
            adherence_changed = False
            drug_changed = False
            hcp_changed = False
            gpo_changed = False
            row_changes = {}
            operations = []

            # 1. Map Medication Adherence %
            adherence_status, adherence_operation, adherence_changed = map_medication_adherence(row['Medication Adherence %'])
            if adherence_changed:
                row_changes['Medication Adherence Status'] = True
                operations.append(adherence_operation)

            # 2. Map Drug Name using GenAI
            mapped_drug, drug_operation, drug_changed = map_drug_name(row['Drug Name'], veeva_product_names)
            if drug_changed:
                row_changes['Mapped Drug Name'] = True
                operations.append(drug_operation)

            # 3. Map HCP ID using GenAI (strictly mapping to 5-digit Veeva IDs)
            corrected_hcp_id, hcp_operation, hcp_changed = map_hcp_id(row['HCP ID'], veeva_hcp_ids)
            if hcp_changed:
                row_changes['Corrected HCP ID'] = True
                operations.append(hcp_operation)

            # 4. Fill Missing GPO Affiliation
            if pd.isna(row['GPO Affiliation']):
                filled_gpo_affiliation = veeva_hcp_to_gpo.get(corrected_hcp_id, "Unknown")
                row_changes['GPO Affiliation'] = True
                operations.append(f'GPO Affiliation: Mapped missing GPO Affiliation for "{row["HCP ID"]}"')
            else:
                filled_gpo_affiliation = row['GPO Affiliation']

            # Store cleaned row and change information
            cleaned_row = {
                'Row #': f'Row {index + 1}',
                'Medication Adherence Status': adherence_status,
                'Mapped Drug Name': mapped_drug,
                'Corrected HCP ID': corrected_hcp_id,
                'GPO Affiliation': filled_gpo_affiliation
            }
            xponent_cleaned_data.append(cleaned_row)
            changed_rows.append(row_changes)

            # Record operations performed
            operations_performed.append(", ".join(operations) if operations else "No Changes")

        # Add the operations column to the DataFrame
        cleaned_df = pd.DataFrame(xponent_cleaned_data)
        cleaned_df['Operations Performed'] = operations_performed

        print("Semantic mapping completed successfully")

        # Save the output with highlights to Excel
        save_to_excel_with_highlight(cleaned_df, output_file_name, changed_rows)

    except Exception as e:
        print(f"Error occurred during semantic mapping: {e}")
        raise
